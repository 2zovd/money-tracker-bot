"""Generate the budget-tracker spreadsheet (.xlsx) from bot/categories.py.

Run: python tracker/build_tracker.py tracker/budget-tracker.xlsx
Then upload it to Google Drive and open it as a Google Sheet (see docs/SETUP.md).
Categories come from bot/categories.py — edit them there, not here."""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from bot import config
from bot.categories import CATEGORY_TABLE, GROUPS

SYM = config.CURRENCY_SYMBOL
CODE = config.CURRENCY_CODE

wb = Workbook()
ARIAL = "Arial"
def F(sz=10, b=False, color="000000"): return Font(name=ARIAL, size=sz, bold=b, color=color)
title_font=F(14,True,"FFFFFF"); hdr_font=F(10,True,"FFFFFF"); sect_font=F(11,True,"1F3864")
input_font=F(10,False,"0000FF"); note_font=F(9,False,"808080")
title_fill=PatternFill("solid",fgColor="1F3864"); hdr_fill=PatternFill("solid",fgColor="2E5496")
sect_fill=PatternFill("solid",fgColor="D6E0F0"); input_fill=PatternFill("solid",fgColor="FFF2CC")
warn_fill=PatternFill("solid",fgColor="FCE4D6"); good_fill=PatternFill("solid",fgColor="E2EFDA")
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
EUR=f'{SYM}#,##0;[Red]({SYM}#,##0);"-"'; EUR2=f'{SYM}#,##0.00;[Red]({SYM}#,##0.00);"-"'; PCT='0%'

def setc(ws,cell,val,font=None,fill=None,fmt=None,align=None,bd=False):
    c=ws[cell]; c.value=val
    if font:c.font=font
    if fill:c.fill=fill
    if fmt:c.number_format=fmt
    if align:c.alignment=Alignment(horizontal=align,vertical="center")
    if bd:c.border=border
    return c

# Category reference: (name, group, essential, plan, currency, note) — from bot/categories.py
CATS = [(name, grp, "yes" if ess else "no", plan, cur, note)
        for name, grp, ess, plan, cur, note in CATEGORY_TABLE]
N=len(CATS)
LAST=3+N        # last category row
TOTAL=4+N       # TOTAL row in "Expenses"
ESS=TOTAL+2     # "Essential / mo" row

# ============ DASHBOARD
d=wb.active; d.title="Dashboard"; d.sheet_view.showGridLines=False
for col,w in {"A":34,"B":15,"C":3,"D":30,"E":14}.items(): d.column_dimensions[col].width=w
d.merge_cells("A1:E1"); setc(d,"A1","BUDGET TRACKER — DASHBOARD",title_font,title_fill,align="left"); d.row_dimensions[1].height=26
setc(d,"A3","SETTINGS / RATES",sect_font,sect_fill); d.merge_cells("A3:B3")
setc(d,"A4",f"USD → {CODE} rate"); setc(d,"B4",config.USD_RATE,input_font,input_fill,"0.0000",bd=True)
setc(d,"A5",f"GBP → {CODE} rate"); setc(d,"B5",config.GBP_RATE,input_font,input_fill,"0.0000",bd=True)
setc(d,"A6","Report month (YYYY-MM)"); setc(d,'B6','=TEXT(TODAY(),"YYYY-MM")',input_font,input_fill,align="center",bd=True)
setc(d,"A7","Savings goal / month"); setc(d,"B7",800,input_font,input_fill,EUR,bd=True)
setc(d,"A9","CASH FLOW / MONTH",sect_font,sect_fill); d.merge_cells("A9:B9")
flow=[("A10","Income (net)","B10","=Income!B7",EUR,F(10,True)),
 ("A11","Expenses — plan","B11",f"=Expenses!F{TOTAL}",EUR,F(10)),
 ("A12","Expenses — actual (report month)","B12","=SUMIF(Journal!I:I,$B$6,Journal!C:C)",EUR,F(10)),
 ("A13","Essential expenses","B13",f'=SUMIF(Expenses!C4:C{LAST},"yes",Expenses!F4:F{LAST})',EUR,F(10)),
 ("A14","Surplus — plan","B14","=B10-B11",EUR,F(10,True)),
 ("A15","Surplus — actual","B15","=B10-B12",EUR,F(10,True))]
for a,al,b,bl,fmt,ft in flow:
    setc(d,a,al,ft); setc(d,b,bl,ft,None,fmt,bd=True)
d["B14"].fill=good_fill; d["B15"].fill=good_fill
# emergency fund
setc(d,"D3","EMERGENCY FUND",sect_font,sect_fill); d.merge_cells("D3:E3")
pod=[("D4","Current buffer","E4",'=SUMIF(Assets!E:E,"Buffer",Assets!D:D)',EUR),
 ("D5","Goal (6 months essential)","E5","=B13*6",EUR),
 ("D6","Filled, %","E6","=IF(E5=0,0,E4/E5)",PCT),
 ("D7","Left to goal","E7","=MAX(E5-E4,0)",EUR),
 ("D8","Months to goal","E8",'=IF($B$7=0,"—",ROUNDUP(E7/$B$7,0))',"0")]
for a,al,b,bl,fmt in pod:
    setc(d,a,al); setc(d,b,bl,F(10,True),None,fmt,bd=True)
d["E6"].fill=warn_fill
setc(d,"D10","ASSETS (liquid)",sect_font,sect_fill); d.merge_cells("D10:E10")
act=[("D11","Buffer (core)","E11",'=SUMIF(Assets!E:E,"Buffer",Assets!D:D)',EUR),
 ("D12","Risk","E12",'=SUMIF(Assets!E:E,"Risk",Assets!D:D)',EUR),
 ("D13","TOTAL assets","E13","=SUM(Assets!D4:D7)",EUR)]
for a,al,b,bl,fmt in act:
    setc(d,a,al); setc(d,b,bl,F(10,True),None,fmt,bd=True)
d["E13"].fill=good_fill
# group rollup (actual)
setc(d,"A17","GROUP ROLLUP — ACTUAL (report month)",sect_font,sect_fill); d.merge_cells("A17:B17")
r=18
for g in GROUPS:
    setc(d,f"A{r}",g,F(10),bd=True)
    setc(d,f"B{r}",f'=SUMIFS(Journal!$C:$C,Journal!$J:$J,A{r},Journal!$I:$I,$B$6)',F(10),None,EUR,bd=True)
    r+=1
setc(d,f"A{r}","TOTAL actual",F(10,True),good_fill,bd=True)
setc(d,f"B{r}",f"=SUM(B18:B{r-1})",F(10,True),good_fill,EUR,bd=True)
setc(d,"D17","Yellow cells — edit them. Categories/groups live on the Reference sheet.",note_font)
d.merge_cells("D17:E24"); d["D17"].alignment=Alignment(wrap_text=True,vertical="top")

# ============ REFERENCE
sp=wb.create_sheet("Reference"); sp.sheet_view.showGridLines=False
for col,w in {"A":26,"B":16,"C":13,"D":40}.items(): sp.column_dimensions[col].width=w
sp.merge_cells("A1:D1"); setc(sp,"A1","CATEGORY REFERENCE",title_font,title_fill,align="left")
setc(sp,"A2","Category → Group → Essential. The bot and planning read categories from here.",note_font); sp.merge_cells("A2:D2")
for i,h in enumerate(["Category","Group","Essential","Note"]):
    setc(sp,f"{chr(65+i)}3",h,hdr_font,hdr_fill,bd=True,align="center")
r=4
for name,grp,ess,plan,cur,note in CATS:
    setc(sp,f"A{r}",name,F(10),bd=True)
    setc(sp,f"B{r}",grp,F(10),bd=True)
    setc(sp,f"C{r}",ess,F(10),bd=True,align="center")
    setc(sp,f"D{r}",note,note_font,bd=True)
    r+=1

# ============ INCOME (plan)
inc=wb.create_sheet("Income"); inc.sheet_view.showGridLines=False
for col,w in {"A":32,"B":14,"C":42}.items(): inc.column_dimensions[col].width=w
inc.merge_cells("A1:C1"); setc(inc,"A1",f"INCOME ({SYM}/mo, net)",title_font,title_fill,align="left")
for i,h in enumerate(["Source",f"{SYM}/mo","Note"]): setc(inc,f"{chr(65+i)}3",h,hdr_font,hdr_fill,bd=True,align="center")
for i,(name,val,note) in enumerate([("Salary (net)",3000,""),("Reimbursements",0,""),("Bonus (avg/mo)",0,"upside, plan=0")]):
    r=4+i; setc(inc,f"A{r}",name,F(10),bd=True); setc(inc,f"B{r}",val,input_font,input_fill,EUR,bd=True); setc(inc,f"C{r}",note,note_font,bd=True)
setc(inc,"A7","TOTAL income",F(10,True),good_fill,bd=True); setc(inc,"B7","=SUM(B4:B6)",F(10,True),good_fill,EUR,bd=True); setc(inc,"C7","",None,bd=True)

# ============ EXPENSES (plan)
ex=wb.create_sheet("Expenses"); ex.sheet_view.showGridLines=False
widths={"A":24,"B":15,"C":10,"D":13,"E":8,"F":12,"G":12,"H":13,"I":32}
for col,w in widths.items(): ex.column_dimensions[col].width=w
ex.merge_cells("A1:I1"); setc(ex,"A1","EXPENSES — PLAN vs ACTUAL",title_font,title_fill,align="left")
setc(ex,"A2","Group and Essential come from Reference. Actual comes from the Journal for the report month.",note_font); ex.merge_cells("A2:I2")
for i,h in enumerate(["Category","Group","Essential","Amount (orig.)","Currency",f"{SYM} / mo","Actual (mo)","Diff","Note"]):
    setc(ex,f"{chr(65+i)}3",h,hdr_font,hdr_fill,bd=True,align="center")
r=4
for name,grp,ess,plan,cur,note in CATS:
    setc(ex,f"A{r}",name,F(10),bd=True)
    setc(ex,f"B{r}",f'=IFERROR(VLOOKUP(A{r},Reference!$A:$C,2,0),"")',F(10),None,None,bd=True)
    setc(ex,f"C{r}",f'=IFERROR(VLOOKUP(A{r},Reference!$A:$C,3,0),"")',F(10),None,None,bd=True,align="center")
    setc(ex,f"D{r}",plan,input_font,input_fill,EUR2,bd=True)
    setc(ex,f"E{r}",cur,input_font,input_fill,None,bd=True,align="center")
    setc(ex,f"F{r}",f'=D{r}*IF(E{r}="{CODE}",1,IF(E{r}="USD",Dashboard!$B$4,IF(E{r}="GBP",Dashboard!$B$5,1)))',F(10),None,EUR,bd=True)
    setc(ex,f"G{r}",f'=SUMIFS(Journal!$C:$C,Journal!$B:$B,A{r},Journal!$I:$I,Dashboard!$B$6)',F(10),None,EUR,bd=True)
    setc(ex,f"H{r}",f'=G{r}-F{r}',F(10),None,EUR,bd=True)
    setc(ex,f"I{r}",note,note_font,bd=True)
    r+=1
setc(ex,f"A{TOTAL}","TOTAL",F(10,True),good_fill,bd=True)
for col in ["B","C","D","E"]: setc(ex,f"{col}{TOTAL}","",None,good_fill,bd=True)
setc(ex,f"F{TOTAL}",f"=SUM(F4:F{LAST})",F(10,True),good_fill,EUR,bd=True)
setc(ex,f"G{TOTAL}",f"=SUM(G4:G{LAST})",F(10,True),good_fill,EUR,bd=True)
setc(ex,f"H{TOTAL}",f"=SUM(H4:H{LAST})",F(10,True),good_fill,EUR,bd=True)
setc(ex,f"I{TOTAL}","",None,good_fill,bd=True)
setc(ex,f"A{ESS}","Essential / mo",F(10,True)); setc(ex,f"F{ESS}",f'=SUMIF(C4:C{LAST},"yes",F4:F{LAST})',F(10,True),good_fill,EUR,bd=True)

# ============ JOURNAL
jr=wb.create_sheet("Journal"); jr.sheet_view.showGridLines=False
widths={"A":12,"B":22,"C":11,"D":16,"E":9,"F":8,"G":10,"H":26,"I":10,"J":15,"K":9}
for col,w in widths.items(): jr.column_dimensions[col].width=w
jr.merge_cells("A1:K1"); setc(jr,"A1","EXPENSE JOURNAL — add each expense (or via the bot)",title_font,title_fill,align="left")
heads=["Date","Category","Amount","Place","Method","Liters","Price/L","Note","Month","Group","Essential"]
for i,h in enumerate(heads): setc(jr,f"{chr(65+i)}3",h,hdr_font,hdr_fill,bd=True,align="center")
def jrow(r):
    for col in ["A","B","C","D","E","F","H"]:
        setc(jr,f"{col}{r}","",F(10),bd=True)
    jr[f"C{r}"].number_format=EUR2; jr[f"A{r}"].alignment=Alignment(horizontal="center")
    jr[f"F{r}"].number_format='0.00'
    setc(jr,f"G{r}",f'=IF(N(F{r})=0,"",C{r}/F{r})',F(10),None,EUR2,bd=True)
    setc(jr,f"I{r}",f'=IF(A{r}="","",TEXT(A{r},"YYYY-MM"))',note_font,None,None,bd=True,align="center")
    setc(jr,f"J{r}",f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},Reference!$A:$C,2,0),""))',note_font,None,None,bd=True)
    setc(jr,f"K{r}",f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},Reference!$A:$C,3,0),""))',note_font,None,None,bd=True,align="center")
# example row 4
setc(jr,"A4","2026-06-15",F(10),bd=True,align="center")
setc(jr,"B4","Groceries",F(10),bd=True)
setc(jr,"C4",23.5,F(10),None,EUR2,bd=True)
setc(jr,"D4","Store A",F(10),bd=True)
setc(jr,"E4","card",F(10),bd=True,align="center")
setc(jr,"F4","",F(10),bd=True); jr["F4"].number_format='0.00'
setc(jr,"H4","example — you can delete",note_font,bd=True)
setc(jr,"G4",'=IF(N(F4)=0,"",C4/F4)',F(10),None,EUR2,bd=True)
setc(jr,"I4",'=IF(A4="","",TEXT(A4,"YYYY-MM"))',note_font,None,None,bd=True,align="center")
setc(jr,"J4",'=IF(B4="","",IFERROR(VLOOKUP(B4,Reference!$A:$C,2,0),""))',note_font,None,None,bd=True)
setc(jr,"K4",'=IF(B4="","",IFERROR(VLOOKUP(B4,Reference!$A:$C,3,0),""))',note_font,None,None,bd=True,align="center")
for r in range(5,305): jrow(r)
dv_cat=DataValidation(type="list",formula1=f"=Reference!$A$4:$A${LAST}",allow_blank=True)
jr.add_data_validation(dv_cat); dv_cat.add("B4:B304")
dv_way=DataValidation(type="list",formula1='"cash,card"',allow_blank=True)
jr.add_data_validation(dv_way); dv_way.add("E4:E304")

# ============ ASSETS
av=wb.create_sheet("Assets"); av.sheet_view.showGridLines=False
widths={"A":24,"B":13,"C":8,"D":13,"E":13,"F":30}
for col,w in widths.items(): av.column_dimensions[col].width=w
av.merge_cells("A1:F1"); setc(av,"A1","ASSETS / SAVINGS",title_font,title_fill,align="left")
for i,h in enumerate(["Asset","Amount (orig.)","Currency",SYM,"Type","Note"]): setc(av,f"{chr(65+i)}3",h,hdr_font,hdr_fill,bd=True,align="center")
for i,(name,amt,cur,typ,note) in enumerate([("Cash",700,"EUR","Buffer","set aside"),("Savings",200,"EUR","Buffer",""),("Investments",0,"EUR","Risk","")]):
    r=4+i
    setc(av,f"A{r}",name,F(10),bd=True); setc(av,f"B{r}",amt,input_font,input_fill,EUR2,bd=True)
    setc(av,f"C{r}",cur,input_font,input_fill,None,bd=True,align="center")
    setc(av,f"D{r}",f'=B{r}*IF(C{r}="{CODE}",1,IF(C{r}="USD",Dashboard!$B$4,IF(C{r}="GBP",Dashboard!$B$5,1)))',F(10),None,EUR,bd=True)
    setc(av,f"E{r}",typ,F(10),bd=True,align="center"); setc(av,f"F{r}",note,note_font,bd=True)
setc(av,"A8","TOTAL assets",F(10,True),good_fill,bd=True)
for col in ["B","C"]: setc(av,f"{col}8","",None,good_fill,bd=True)
setc(av,"D8","=SUM(D4:D7)",F(10,True),good_fill,EUR,bd=True)
for col in ["E","F"]: setc(av,f"{col}8","",None,good_fill,bd=True)

out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), "budget-tracker.xlsx")
wb.save(out)
print("saved:", out)
